# -*- coding: utf-8 -*-
"""Modelo relacional derivado SOLO de network/dto (sin tablas inventadas)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent / "modelo_relacional_desde_dtos.xlsx"

COLS = [
    "orden",
    "atributo",
    "tipo_dto",
    "json_api",
    "pk",
    "fk_referencia",
    "nullable",
    "default_dto",
    "notas",
]

HEADER_FILL = PatternFill("solid", fgColor="2E5090")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(italic=True, size=10, color="555555")
WARN_FILL = PatternFill("solid", fgColor="FFF3CD")


def a(orden, nombre, tipo, json_api="", pk="", fk="", nullable="Sí", default="", notas=""):
    return [orden, nombre, tipo, json_api, pk, fk, nullable, default, notas]


# --- Entidades = data class del API (persistibles) ---
ENTITIES = {
    "Company": [
        a(1, "ruc", "String", "ruc", pk="PK", nullable="No", notas="DTO: Company.kt · También embebido en Usuario"),
        a(2, "nombre", "String", "nombre", nullable="No"),
        a(3, "rutaFirma", "String?", "ruta_firma"),
        a(4, "rutaLogo", "String?", "ruta_logo"),
        a(5, "nameLogo", "String?", "name_logo"),
        a(6, "direccion", "String?", "direccion"),
        a(7, "telefono", "String?", "telefono"),
    ],
    "Usuario": [
        a(1, "email", "String", "email", pk="PK", nullable="No", notas="DTO: Usuario.kt · Respuesta POST /auth/login"),
        a(2, "contrasena", "String", "contrasena", nullable="No", notas="En API login; en BD debe ser hash"),
        a(3, "token", "String?", "token"),
        a(4, "refreshToken", "String?", "refresh_token"),
        a(5, "lastUpdated", "Long", "last_updated", nullable="No"),
        a(6, "estado", "EstadoUsuario", "estado", nullable="No", notas="ACTIVO | DISABLED | DELETED"),
        a(7, "company", "Company?", "company", notas="Objeto anidado → FK lógica company.ruc"),
    ],
    "Cliente": [
        a(1, "id", "String", "id", pk="PK", nullable="No"),
        a(2, "companyRuc", "String", "company_ruc", fk="Company.ruc", nullable="No"),
        a(3, "tipoDoc", "String", "tipo_doc", nullable="No", notas="1=DNI, 6=RUC"),
        a(4, "numeroDoc", "String", "numero_doc", nullable="No"),
        a(5, "razonSocial", "String", "razon_social", nullable="No"),
        a(6, "direccion", "String?", "direccion"),
        a(7, "telefono", "String?", "telefono"),
        a(8, "activo", "Boolean", "activo", nullable="No", default="true"),
    ],
    "CatalogItem": [
        a(1, "id", "String", "id", pk="PK", nullable="No"),
        a(2, "companyRuc", "String", "company_ruc", fk="Company.ruc", nullable="No"),
        a(3, "kind", "String", "kind", nullable="No", notas="PRODUCT | SERVICE → CatalogItemKind"),
        a(4, "codigo", "String?", "codigo"),
        a(5, "nombre", "String", "nombre", nullable="No"),
        a(6, "descripcion", "String?", "descripcion"),
        a(7, "unidad", "String", "unidad", nullable="No", notas="NIU, MTR, KGM, LTR…"),
        a(8, "precioUnitario", "Double", "precio_unitario", nullable="No"),
        a(9, "afectacionIgv", "String", "afectacion_igv", nullable="No", default="10"),
        a(10, "activo", "Boolean", "activo", nullable="No", default="true"),
        a(11, "manejaStock", "Boolean", "maneja_stock", nullable="No", default="false"),
        a(12, "manejaSerie", "Boolean", "maneja_serie", nullable="No", default="false"),
        a(13, "stockActual", "Double?", "stock_actual", notas="Solo lectura; query ?almacen_id= en GET catálogo"),
        a(14, "duracionMinutos", "Int?", "duracion_minutos", notas="Servicios"),
    ],
    "Almacen": [
        a(1, "id", "String", "id", pk="PK", nullable="No"),
        a(2, "companyRuc", "String", "company_ruc", fk="Company.ruc", nullable="No"),
        a(3, "codigo", "String", "codigo", nullable="No"),
        a(4, "nombre", "String", "nombre", nullable="No"),
        a(5, "direccion", "String?", "direccion"),
        a(6, "activo", "Boolean", "activo", nullable="No", default="true"),
    ],
    "ProductoSerie": [
        a(1, "id", "String", "id", pk="PK", nullable="No"),
        a(2, "companyRuc", "String", "company_ruc", fk="Company.ruc", nullable="No"),
        a(3, "catalogItemId", "String", "catalog_item_id", fk="CatalogItem.id", nullable="No"),
        a(4, "numeroSerie", "String", "numero_serie", nullable="No"),
        a(5, "almacenId", "String?", "almacen_id", fk="Almacen.id"),
        a(6, "estado", "ProductoSerieEstado", "estado", nullable="No", notas="DISPONIBLE|RESERVADO|VENDIDO|ENTREGADO|BAJA"),
        a(7, "comprobanteId", "String?", "comprobante_id", fk="Comprobante.id"),
        a(8, "entregaId", "String?", "entrega_id", fk="Movimiento.id", notas="Salida/entrega que despachó la unidad"),
    ],
    "Movimiento": [
        a(1, "id", "String", "id", pk="PK", nullable="No"),
        a(2, "companyRuc", "String", "company_ruc", fk="Company.ruc", nullable="No"),
        a(3, "almacenId", "String", "almacen_id", fk="Almacen.id", nullable="No", notas="Origen; en ENTRADA = destino físico"),
        a(4, "tipo", "MovimientoTipo", "tipo", nullable="No", notas="ENTRADA | SALIDA | AJUSTE"),
        a(5, "lineas", "List<MovimientoLinea>", "lineas", nullable="No", notas="1:N → entidad MovimientoLinea"),
        a(6, "fecha", "String", "fecha", nullable="No", notas="ISO-8601"),
        a(7, "observaciones", "String?", "observaciones"),
        a(8, "referenciaTipo", "String?", "referencia_tipo"),
        a(9, "referenciaId", "String?", "referencia_id"),
        a(10, "numero", "String?", "numero", notas="Ej. ENT-00001 (salidas)"),
        a(11, "almacenDestinoId", "String?", "almacen_destino_id", fk="Almacen.id", notas="Traslado interno"),
        a(12, "estado", "MovimientoEstado?", "estado", notas="BORRADOR | DESPACHADA | ANULADA"),
        a(13, "comprobanteId", "String?", "comprobante_id", fk="Comprobante.id"),
        a(14, "cliente", "MovimientoCliente?", "cliente", notas="Objeto anidado (salida a cliente)"),
        a(15, "fechaDespacho", "String?", "fecha_despacho"),
    ],
    "MovimientoLinea": [
        a(1, "id", "String?", "id", pk="PK*", notas="*Opcional en JSON; PK en BD = id o compuesta"),
        a(2, "movimientoId", "—", "—", fk="Movimiento.id", nullable="No", notas="Implícito: pertenece a lineas[]"),
        a(3, "cantidad", "Double", "cantidad", nullable="No"),
        a(4, "catalogItemId", "String", "catalog_item_id", fk="CatalogItem.id", nullable="No", notas="vía catalogItemIdSinSerie o serie.catalogItemId"),
        a(5, "serie", "ProductoSerie?", "serie", fk="ProductoSerie.id", notas="GET: unidad embebida"),
        a(6, "numerosSerie", "List<String>?", "series", notas="POST ingreso: SN en texto"),
        a(7, "serieIds", "List<String>?", "serie_ids", notas="POST salida: IDs existentes"),
    ],
    "MovimientoCliente": [
        a(1, "—", "—", notas="No es tabla aparte en DTO; columnas dentro de Movimiento"),
        a(2, "tipoDoc", "String", "cliente.tipo_doc", nullable="No"),
        a(3, "numeroDoc", "String", "cliente.numero_doc", nullable="No"),
        a(4, "razonSocial", "String?", "cliente.razon_social"),
    ],
    "Comprobante": [
        a(1, "id", "String", "id", pk="PK", nullable="No"),
        a(2, "companyRuc", "String", "company_ruc", fk="Company.ruc", nullable="No"),
        a(3, "tipo", "String", "tipo", nullable="No", notas="FACTURA|BOLETA|NOTA_*|GUIA_EMISION"),
        a(4, "serie", "String", "serie", nullable="No"),
        a(5, "numero", "String", "numero", nullable="No"),
        a(6, "estado", "ComprobanteEstado", "estado", nullable="No"),
        a(7, "receptor", "ComprobanteReceptor", "receptor", nullable="No", notas="Objeto anidado"),
        a(8, "lineas", "List<LineaComprobante>", "lineas", nullable="No", notas="1:N → LineaComprobante"),
        a(9, "totales", "ComprobanteTotales", "totales", nullable="No", notas="Objeto anidado"),
        a(10, "documentoAfectado", "ComprobanteReferencia?", "documento_afectado", notas="Notas crédito/débito"),
        a(11, "facturas", "List<ComprobanteReferencia>?", "facturas", notas="Guía de emisión → N facturas"),
        a(12, "motivoNota", "String?", "motivo_nota"),
        a(13, "observaciones", "String?", "observaciones"),
        a(14, "fechaEmision", "String?", "fecha_emision"),
        a(15, "cdrEstado", "String?", "cdr_estado"),
        a(16, "pdfUrl", "String?", "pdf_url"),
        a(17, "cdrZipUrl", "String?", "cdr_zip_url"),
    ],
    "ComprobanteReceptor": [
        a(1, "—", "—", notas="Embebido en Comprobante (snapshot; no FK a Cliente)"),
        a(2, "tipoDoc", "String", "receptor.tipo_doc", nullable="No"),
        a(3, "numeroDoc", "String", "receptor.numero_doc", nullable="No"),
        a(4, "razonSocial", "String", "receptor.razon_social", nullable="No"),
    ],
    "ComprobanteTotales": [
        a(1, "—", "—", notas="Embebido en Comprobante"),
        a(2, "subtotal", "Double", "totales.subtotal", nullable="No"),
        a(3, "igv", "Double", "totales.igv", nullable="No"),
        a(4, "total", "Double", "totales.total", nullable="No"),
        a(5, "moneda", "String", "totales.moneda", nullable="No", default="PEN"),
    ],
    "LineaComprobante": [
        a(1, "id", "String?", "id", pk="PK*"),
        a(2, "comprobanteId", "—", "—", fk="Comprobante.id", nullable="No", notas="Implícito en lineas[]"),
        a(3, "catalogItemId", "String?", "catalog_item_id", fk="CatalogItem.id"),
        a(4, "descripcion", "String", "descripcion", nullable="No"),
        a(5, "cantidad", "Double", "cantidad", nullable="No"),
        a(6, "unidad", "String", "unidad", nullable="No"),
        a(7, "precioUnitario", "Double", "precio_unitario", nullable="No"),
        a(8, "afectacionIgv", "String", "afectacion_igv", nullable="No", default="10"),
        a(9, "subtotal", "Double", "subtotal", nullable="No"),
        a(10, "igv", "Double", "igv", nullable="No"),
        a(11, "total", "Double", "total", nullable="No"),
        a(12, "series", "List<ProductoSerie>?", "series", notas="Unidades serializadas en la línea"),
    ],
    "ComprobanteReferencia": [
        a(1, "—", "—", notas="Puede repetirse: documento_afectado (0..1) o facturas[] (0..N)"),
        a(2, "comprobanteId", "—", fk="Comprobante.id (guía o nota)", nullable="No"),
        a(3, "tipo", "String", "tipo", nullable="No"),
        a(4, "serie", "String", "serie", nullable="No"),
        a(5, "numero", "String", "numero", nullable="No"),
        a(6, "fechaEmision", "String?", "fecha_emision"),
    ],
    "EventoVidaProducto": [
        a(1, "id", "String", "id", pk="PK", nullable="No", notas="DTO presente; NO hay ruta en ApiService"),
        a(2, "catalogItemId", "String", "catalogItemId", fk="CatalogItem.id", nullable="No"),
        a(3, "nombreProducto", "String", "nombreProducto", nullable="No"),
        a(4, "tipo", "TipoEventoInventario", "tipo", nullable="No", notas="INGRESO|SALIDA|SERIE"),
        a(5, "titulo", "String", "titulo", nullable="No"),
        a(6, "detalle", "String", "detalle", nullable="No"),
        a(7, "fecha", "String", "fecha", nullable="No"),
        a(8, "movimientoId", "String?", "movimiento_id", fk="Movimiento.id"),
        a(9, "almacenId", "String?", "almacen_id", fk="Almacen.id"),
        a(10, "almacenDestinoId", "String?", "almacen_destino_id", fk="Almacen.id"),
        a(11, "cantidad", "Double?", "cantidad"),
        a(12, "registradoPor", "String?", "registrado_por", notas="No existe en Movimiento DTO"),
        a(13, "observaciones", "String?", "observaciones"),
        a(14, "referenciaTipo", "String?", "referencia_tipo"),
        a(15, "numeroMovimiento", "String?", "numero_movimiento"),
        a(16, "cliente", "MovimientoCliente?", "cliente"),
        a(17, "numeroSerie", "String?", "numero_serie"),
        a(18, "lineasMovimiento", "List<EventoVidaLineaDetalle>?", "lineas_movimiento"),
    ],
    "EventoVidaLineaDetalle": [
        a(1, "catalogItemId", "String", "catalog_item_id", fk="CatalogItem.id", nullable="No"),
        a(2, "nombreProducto", "String", "nombreProducto", nullable="No"),
        a(3, "cantidad", "Double", "cantidad", nullable="No"),
        a(4, "unidad", "String", "unidad", nullable="No", default="NIU"),
        a(5, "numerosSerie", "List<String>?", "numeros_serie"),
        a(6, "manejaSerie", "Boolean", "maneja_serie", nullable="No", default="false"),
    ],
}

RELACIONES = [
    ("Company", "1", "N", "Cliente", "company_ruc → ruc", "GET/POST …/clientes"),
    ("Company", "1", "N", "CatalogItem", "company_ruc → ruc", "GET …/catalogo"),
    ("Company", "1", "N", "Almacen", "company_ruc → ruc", "GET/POST …/almacenes"),
    ("Company", "1", "N", "Comprobante", "company_ruc → ruc", "GET/POST …/comprobantes · GET …/compras (mismo DTO)"),
    ("Company", "1", "N", "Movimiento", "company_ruc → ruc", "entradas · movimientos · entregas"),
    ("Company", "1", "N", "ProductoSerie", "company_ruc → ruc", "series-disponibles"),
    ("Usuario", "N", "0..1", "Company", "company (anidado)", "login response"),
    ("CatalogItem", "1", "N", "ProductoSerie", "catalog_item_id → id", ""),
    ("CatalogItem", "1", "N", "MovimientoLinea", "catalog_item_id", ""),
    ("CatalogItem", "1", "N", "LineaComprobante", "catalog_item_id (opcional)", "snapshot"),
    ("Almacen", "1", "N", "Movimiento", "almacen_id", "origen"),
    ("Almacen", "1", "N", "Movimiento", "almacen_destino_id", "traslado (opcional)"),
    ("Almacen", "1", "N", "ProductoSerie", "almacen_id (opcional)", ""),
    ("Movimiento", "1", "N", "MovimientoLinea", "lineas[]", "compuesto en JSON"),
    ("Movimiento", "N", "0..1", "Comprobante", "comprobante_id", "salida vinculada a venta"),
    ("Movimiento", "1", "0..1", "MovimientoCliente", "cliente (anidado)", "solo salida"),
    ("Comprobante", "1", "N", "LineaComprobante", "lineas[]", ""),
    ("Comprobante", "1", "0..1", "ComprobanteReferencia", "documento_afectado", "notas"),
    ("Comprobante", "1", "N", "ComprobanteReferencia", "facturas[]", "guía emisión"),
    ("Comprobante", "1", "N", "ProductoSerie", "comprobante_id", "series vendidas"),
    ("Movimiento", "1", "N", "ProductoSerie", "entrega_id", "series entregadas"),
    ("Movimiento", "1", "N", "EventoVidaProducto", "movimiento_id", "vista/UI; armado en app"),
    ("MovimientoLinea", "N", "0..1", "ProductoSerie", "serie | serie_ids", "varias formas en JSON"),
    ("LineaComprobante", "N", "0..N", "ProductoSerie", "series[]", "embebido en línea"),
    ("Cliente", "—", "—", "Comprobante", "sin FK", "receptor es snapshot, no cliente.id"),
]

REQUEST_DTOS = [
    ("LoginRequest", "email, pin", "POST /auth/login"),
    ("CrearClienteRequest", "tipo_doc, numero_doc, razon_social, direccion?, telefono?", "POST clientes"),
    ("CrearAlmacenRequest", "codigo, nombre, direccion?", "POST almacenes"),
    ("EmitirComprobanteRequest", "company_ruc, tipo, receptor, lineas[], documento_afectado?, facturas?, motivo_nota?, observaciones?", "POST comprobantes"),
    ("EmitirLineaRequest", "catalog_item_id, cantidad, serie_ids?", "dentro de emitir"),
    ("RegistrarEntradaRequest", "company_ruc, almacen_id, lineas[], observaciones?", "POST inventario/entradas"),
    ("RegistrarSalidaRequest", "company_ruc, almacen_id, almacen_destino_id?, comprobante_id?, lineas[], cliente?", "POST entregas"),
    ("RegistrarMovimientoLineaRequest", "catalog_item_id, cantidad, series?, serie_ids?", "líneas entrada/salida"),
]

UI_ONLY = [
    ("LineaCatalogoItem", "Pantalla emisión/ingreso/salida; NO va al API tal cual", "Mapea a EmitirLineaRequest / RegistrarMovimientoLineaRequest"),
    ("TotalesComprobante", "Calculado en UI desde LineaCatalogoItem", "API devuelve ComprobanteTotales"),
    ("CatalogItem.manejaInventario", "Propiedad calculada: esProducto && manejaStock", "InventarioExtensions.kt"),
]

ENUMS = [
    ("EstadoUsuario", "ACTIVO, DISABLED, DELETED", "modelos/EstadoUsuario.kt"),
    ("ComprobanteEstado", "BORRADOR, ENVIADO, ACEPTADO, RECHAZADO, ANULADO", "Comprobante.kt"),
    ("ComprobanteTipo", "FACTURA, BOLETA, NOTA_CREDITO, NOTA_DEBITO, GUIA_EMISION", "object ComprobanteTipo"),
    ("MovimientoTipo", "ENTRADA, SALIDA, AJUSTE", "Movimiento.kt"),
    ("MovimientoEstado", "BORRADOR, DESPACHADA, ANULADA", "Movimiento.kt"),
    ("ProductoSerieEstado", "DISPONIBLE, RESERVADO, VENDIDO, ENTREGADO, BAJA", "ProductoSerie.kt"),
    ("CatalogItemKind", "PRODUCT, SERVICE", "kind en JSON"),
    ("TipoEventoInventario", "INGRESO, SALIDA, SERIE", "EventoVidaProducto.kt"),
]


def style_header(ws, row):
    for col in range(1, len(COLS) + 1):
        c = ws.cell(row=row, column=col, value=COLS[col - 1])
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_entity_sheet(wb, name, rows):
    ws = wb.create_sheet(title=name[:31])
    ws["A1"] = f"Entidad DTO: {name}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Fuente: app/.../network/dto/"
    ws["A2"].font = SUB_FONT
    style_header(ws, 4)
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    widths = [6, 26, 22, 22, 6, 28, 10, 12, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_relaciones(wb):
    ws = wb.create_sheet(title="RELACIONES", index=0)
    ws["A1"] = "Relaciones (modelo relacional desde DTOs)"
    ws["A1"].font = TITLE_FONT
    headers = ["entidad_origen", "card_origen", "card_destino", "entidad_destino", "atributo_enlace", "api"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, row in enumerate(RELACIONES, start=4):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 40

    r = 4 + len(RELACIONES) + 2
    ws.cell(row=r, column=1, value="Diagrama lógico (texto)").font = Font(bold=True)
    diagram = [
        "Company (ruc)",
        "  ├──< Cliente",
        "  ├──< CatalogItem ──< ProductoSerie",
        "  ├──< Almacen",
        "  ├──< Comprobante ──< LineaComprobante",
        "  │       ├── (embebido) ComprobanteReceptor, ComprobanteTotales",
        "  │       └──< ComprobanteReferencia  [documento_afectado | facturas[]]",
        "  └──< Movimiento ──< MovimientoLinea",
        "          ├── (embebido) MovimientoCliente",
        "          └──> Comprobante?  comprobante_id",
        "",
        "EventoVidaProducto: vista derivada de Movimiento (sin endpoint propio)",
        "LineaCatalogoItem: solo UI, no persiste en API",
    ]
    for k, line in enumerate(diagram, start=r + 1):
        ws.cell(row=k, column=1, value=line)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + len(diagram), end_column=6)


def write_list_sheet(wb, title, headers, rows, desc=""):
    ws = wb.create_sheet(title=title[:31])
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if desc:
        ws["A2"] = desc
        ws["A2"].font = SUB_FONT
        hr = 4
    else:
        hr = 3
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, row in enumerate(rows, start=hr + 1):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 45


def main():
    wb = Workbook()
    wb.remove(wb.active)
    write_relaciones(wb)
    order = [
        "Company",
        "Usuario",
        "Cliente",
        "CatalogItem",
        "Almacen",
        "ProductoSerie",
        "Movimiento",
        "MovimientoLinea",
        "MovimientoCliente",
        "Comprobante",
        "ComprobanteReceptor",
        "ComprobanteTotales",
        "LineaComprobante",
        "ComprobanteReferencia",
        "EventoVidaProducto",
        "EventoVidaLineaDetalle",
    ]
    for name in order:
        write_entity_sheet(wb, name, ENTITIES[name])
    write_list_sheet(
        wb,
        "REQUEST_body",
        ["dto", "campos", "endpoint"],
        REQUEST_DTOS,
        "Cuerpos POST (no son tablas; mapean a entidades)",
    )
    write_list_sheet(
        wb,
        "SOLO_UI",
        ["nombre", "descripcion", "mapeo"],
        UI_ONLY,
    )
    write_list_sheet(
        wb,
        "ENUMS",
        ["enum", "valores", "archivo"],
        ENUMS,
    )
    wb.save(OUTPUT)
    print(f"OK: {OUTPUT}")


if __name__ == "__main__":
    main()
