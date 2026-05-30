# -*- coding: utf-8 -*-
"""Genera modelo_relacional_backend.xlsx desde el esquema alineado con la app FactApp."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent / "modelo_relacional_backend.xlsx"

# Columnas de cada hoja de tabla
COLS = [
    "orden",
    "atributo",
    "tipo_sql",
    "longitud",
    "pk",
    "fk",
    "nullable",
    "default",
    "unico",
    "indice",
    "campo_json_api",
    "descripcion",
    "notas",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(italic=True, size=10, color="444444")


def attr(
    orden,
    nombre,
    tipo,
    longitud="",
    pk="",
    fk="",
    nullable="Sí",
    default="",
    unico="",
    indice="",
    json_api="",
    desc="",
    notas="",
):
    return [
        orden,
        nombre,
        tipo,
        longitud,
        pk,
        fk,
        nullable,
        default,
        unico,
        indice,
        json_api,
        desc,
        notas,
    ]


TABLES = {
    "00_INDICE": [],  # special sheet
    "empresa": [
        attr(1, "ruc", "VARCHAR", "11", pk="PK", nullable="No", json_api="ruc", desc="RUC emisor (tenant)"),
        attr(2, "nombre", "VARCHAR", "255", nullable="No", json_api="nombre", desc="Razón social"),
        attr(3, "direccion", "VARCHAR", "500", json_api="direccion", desc="Domicilio fiscal"),
        attr(4, "telefono", "VARCHAR", "30", json_api="telefono"),
        attr(5, "ruta_firma", "VARCHAR", "500", json_api="ruta_firma", desc="Ruta certificado .pfx"),
        attr(6, "ruta_logo", "VARCHAR", "500", json_api="ruta_logo"),
        attr(7, "name_logo", "VARCHAR", "255", json_api="name_logo", desc="Nombre archivo logo"),
        attr(8, "plantilla", "VARCHAR", "20", nullable="No", default="GENERAL", json_api="—", desc="GENERAL | RETAIL | SERVICIOS", notas="Room local; API futuro"),
        attr(9, "activo", "BOOLEAN", nullable="No", default="true"),
        attr(10, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(11, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
    ],
    "usuario": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="—", desc="Identificador interno"),
        attr(2, "email", "VARCHAR", "255", unico="Sí", indice="idx_usuario_email", nullable="No", json_api="email", desc="Login (PK lógica en app)"),
        attr(3, "pin_hash", "VARCHAR", "255", nullable="No", json_api="contrasena", desc="Hash PIN 6 dígitos; nunca texto plano"),
        attr(4, "nombre", "VARCHAR", "150", desc="Nombre para mostrar"),
        attr(5, "estado", "VARCHAR", "20", nullable="No", default="ACTIVO", json_api="estado", desc="ACTIVO | DISABLED | DELETED"),
        attr(6, "token_version", "INTEGER", nullable="No", default="0", desc="Invalidar JWT al cambiar PIN"),
        attr(7, "last_updated", "BIGINT", json_api="last_updated", desc="Epoch ms (respuesta login)"),
        attr(8, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(9, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
    ],
    "usuario_empresa": [
        attr(1, "usuario_id", "UUID", pk="PK", fk="usuario.id", nullable="No"),
        attr(2, "company_ruc", "VARCHAR", "11", pk="PK", fk="empresa.ruc", nullable="No"),
        attr(3, "rol", "VARCHAR", "30", nullable="No", default="OPERADOR", desc="ADMIN | OPERADOR | SOLO_LECTURA"),
        attr(4, "es_principal", "BOOLEAN", nullable="No", default="false", desc="Empresa por defecto al login"),
        attr(5, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
    ],
    "sesion_refresh": [
        attr(1, "id", "UUID", pk="PK", nullable="No"),
        attr(2, "usuario_id", "UUID", fk="usuario.id", nullable="No", indice="idx_refresh_usuario"),
        attr(3, "token_hash", "VARCHAR", "255", nullable="No", unico="Sí"),
        attr(4, "expires_at", "TIMESTAMPTZ", nullable="No"),
        attr(5, "revoked_at", "TIMESTAMPTZ"),
        attr(6, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
    ],
    "cliente": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "company_ruc", "VARCHAR", "11", fk="empresa.ruc", nullable="No", indice="idx_cliente_empresa", json_api="company_ruc"),
        attr(3, "tipo_doc", "CHAR", "1", nullable="No", json_api="tipo_doc", desc="1=DNI, 6=RUC (SUNAT)"),
        attr(4, "numero_doc", "VARCHAR", "15", nullable="No", json_api="numero_doc"),
        attr(5, "razon_social", "VARCHAR", "255", nullable="No", json_api="razon_social"),
        attr(6, "direccion", "VARCHAR", "500", json_api="direccion"),
        attr(7, "telefono", "VARCHAR", "30", json_api="telefono"),
        attr(8, "activo", "BOOLEAN", nullable="No", default="true", json_api="activo"),
        attr(9, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(10, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(11, "—", "—", unico="Sí", notas="UNIQUE (company_ruc, tipo_doc, numero_doc)"),
    ],
    "catalog_item": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "company_ruc", "VARCHAR", "11", fk="empresa.ruc", nullable="No", indice="idx_catalog_empresa", json_api="company_ruc"),
        attr(3, "kind", "VARCHAR", "10", nullable="No", json_api="kind", desc="PRODUCT | SERVICE"),
        attr(4, "codigo", "VARCHAR", "50", json_api="codigo", desc="SKU / código interno"),
        attr(5, "nombre", "VARCHAR", "255", nullable="No", json_api="nombre"),
        attr(6, "descripcion", "TEXT", json_api="descripcion"),
        attr(7, "unidad", "VARCHAR", "3", nullable="No", default="NIU", json_api="unidad", desc="NIU, MTR, KGM, LTR, ZZ…"),
        attr(8, "precio_unitario", "DECIMAL", "18,4", nullable="No", default="0", json_api="precio_unitario"),
        attr(9, "afectacion_igv", "CHAR", "2", nullable="No", default="10", json_api="afectacion_igv", desc="Catálogo SUNAT IGV"),
        attr(10, "activo", "BOOLEAN", nullable="No", default="true", json_api="activo"),
        attr(11, "maneja_stock", "BOOLEAN", nullable="No", default="false", json_api="maneja_stock"),
        attr(12, "maneja_serie", "BOOLEAN", nullable="No", default="false", json_api="maneja_serie"),
        attr(13, "duracion_minutos", "INTEGER", json_api="duracion_minutos", desc="Solo servicios"),
        attr(14, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(15, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(16, "stock_actual", "DECIMAL", "18,4", json_api="stock_actual", desc="Calculado por almacén; no persistir en catálogo", notas="Vista/API query almacen_id"),
    ],
    "almacen": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "company_ruc", "VARCHAR", "11", fk="empresa.ruc", nullable="No", indice="idx_almacen_empresa", json_api="company_ruc"),
        attr(3, "codigo", "VARCHAR", "20", nullable="No", json_api="codigo"),
        attr(4, "nombre", "VARCHAR", "150", nullable="No", json_api="nombre"),
        attr(5, "direccion", "VARCHAR", "500", json_api="direccion"),
        attr(6, "activo", "BOOLEAN", nullable="No", default="true", json_api="activo"),
        attr(7, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(8, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(9, "—", "—", unico="Sí", notas="UNIQUE (company_ruc, codigo)"),
    ],
    "stock_almacen": [
        attr(1, "company_ruc", "VARCHAR", "11", pk="PK", fk="empresa.ruc", nullable="No"),
        attr(2, "almacen_id", "UUID", pk="PK", fk="almacen.id", nullable="No"),
        attr(3, "catalog_item_id", "UUID", pk="PK", fk="catalog_item.id", nullable="No"),
        attr(4, "cantidad", "DECIMAL", "18,4", nullable="No", default="0", desc="Stock no serializado"),
        attr(5, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(6, "—", "—", notas="Para maneja_serie: stock = COUNT(producto_serie WHERE estado=DISPONIBLE)"),
    ],
    "producto_serie": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "company_ruc", "VARCHAR", "11", fk="empresa.ruc", nullable="No", json_api="company_ruc"),
        attr(3, "catalog_item_id", "UUID", fk="catalog_item.id", nullable="No", indice="idx_serie_producto", json_api="catalog_item_id"),
        attr(4, "numero_serie", "VARCHAR", "100", nullable="No", json_api="numero_serie"),
        attr(5, "almacen_id", "UUID", fk="almacen.id", json_api="almacen_id"),
        attr(6, "estado", "VARCHAR", "20", nullable="No", default="DISPONIBLE", json_api="estado", desc="DISPONIBLE|RESERVADO|VENDIDO|ENTREGADO|BAJA"),
        attr(7, "comprobante_id", "UUID", fk="comprobante.id", json_api="comprobante_id"),
        attr(8, "entrega_id", "UUID", fk="movimiento.id", json_api="entrega_id", desc="Salida/entrega que despachó la unidad"),
        attr(9, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(10, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(11, "—", "—", unico="Sí", notas="UNIQUE (company_ruc, catalog_item_id, numero_serie)"),
    ],
    "movimiento": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "company_ruc", "VARCHAR", "11", fk="empresa.ruc", nullable="No", indice="idx_mov_empresa_fecha", json_api="company_ruc"),
        attr(3, "almacen_id", "UUID", fk="almacen.id", nullable="No", json_api="almacen_id", desc="Origen (salida) o destino (entrada)"),
        attr(4, "tipo", "VARCHAR", "10", nullable="No", json_api="tipo", desc="ENTRADA | SALIDA | AJUSTE"),
        attr(5, "fecha", "TIMESTAMPTZ", nullable="No", json_api="fecha"),
        attr(6, "observaciones", "TEXT", json_api="observaciones"),
        attr(7, "referencia_tipo", "VARCHAR", "30", json_api="referencia_tipo", desc="COMPRA, COMPROBANTE, AJUSTE…"),
        attr(8, "referencia_id", "UUID", json_api="referencia_id"),
        attr(9, "numero", "VARCHAR", "30", json_api="numero", desc="Correlativo ENT-00001 (salidas)"),
        attr(10, "almacen_destino_id", "UUID", fk="almacen.id", json_api="almacen_destino_id", desc="Traslado interno"),
        attr(11, "estado", "VARCHAR", "20", json_api="estado", desc="BORRADOR | DESPACHADA | ANULADA"),
        attr(12, "comprobante_id", "UUID", fk="comprobante.id", json_api="comprobante_id"),
        attr(13, "cliente_tipo_doc", "CHAR", "1", json_api="cliente.tipo_doc"),
        attr(14, "cliente_numero_doc", "VARCHAR", "15", json_api="cliente.numero_doc"),
        attr(15, "cliente_razon_social", "VARCHAR", "255", json_api="cliente.razon_social"),
        attr(16, "fecha_despacho", "TIMESTAMPTZ", json_api="fecha_despacho"),
        attr(17, "registrado_por", "UUID", fk="usuario.id", desc="Usuario que registró"),
        attr(18, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(19, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
    ],
    "movimiento_linea": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "movimiento_id", "UUID", fk="movimiento.id", nullable="No", indice="idx_mov_linea_mov"),
        attr(3, "catalog_item_id", "UUID", fk="catalog_item.id", nullable="No", json_api="catalog_item_id"),
        attr(4, "cantidad", "DECIMAL", "18,4", nullable="No", json_api="cantidad"),
        attr(5, "orden", "SMALLINT", nullable="No", default="1"),
        attr(6, "—", "—", notas="Series en movimiento_linea_serie o JSON legacy numerosSerie"),
    ],
    "movimiento_linea_serie": [
        attr(1, "movimiento_linea_id", "UUID", pk="PK", fk="movimiento_linea.id", nullable="No"),
        attr(2, "producto_serie_id", "UUID", pk="PK", fk="producto_serie.id", nullable="No"),
        attr(3, "numero_serie_texto", "VARCHAR", "100", desc="Alta en ingreso antes de existir UUID"),
        attr(4, "tipo_registro", "VARCHAR", "10", nullable="No", desc="ID_EXISTENTE | NUEVO_SN"),
    ],
    "comprobante": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "company_ruc", "VARCHAR", "11", fk="empresa.ruc", nullable="No", indice="idx_comp_empresa", json_api="company_ruc"),
        attr(3, "direccion", "VARCHAR", "10", nullable="No", default="EMITIDO", desc="EMITIDO (venta) | RECIBIDO (compra)", notas="GET /compras = RECIBIDO"),
        attr(4, "tipo", "VARCHAR", "20", nullable="No", json_api="tipo", desc="FACTURA|BOLETA|NOTA_CREDITO|NOTA_DEBITO|GUIA_EMISION"),
        attr(5, "serie", "VARCHAR", "10", nullable="No", json_api="serie"),
        attr(6, "numero", "VARCHAR", "20", nullable="No", json_api="numero"),
        attr(7, "estado", "VARCHAR", "20", nullable="No", json_api="estado", desc="BORRADOR|ENVIADO|ACEPTADO|RECHAZADO|ANULADO"),
        attr(8, "receptor_tipo_doc", "CHAR", "1", nullable="No", json_api="receptor.tipo_doc"),
        attr(9, "receptor_numero_doc", "VARCHAR", "15", nullable="No", json_api="receptor.numero_doc"),
        attr(10, "receptor_razon_social", "VARCHAR", "255", nullable="No", json_api="receptor.razon_social"),
        attr(11, "subtotal", "DECIMAL", "18,2", nullable="No", json_api="totales.subtotal"),
        attr(12, "igv", "DECIMAL", "18,2", nullable="No", json_api="totales.igv"),
        attr(13, "total", "DECIMAL", "18,2", nullable="No", json_api="totales.total"),
        attr(14, "moneda", "CHAR", "3", nullable="No", default="PEN", json_api="totales.moneda"),
        attr(15, "motivo_nota", "VARCHAR", "255", json_api="motivo_nota"),
        attr(16, "observaciones", "TEXT", json_api="observaciones"),
        attr(17, "fecha_emision", "TIMESTAMPTZ", json_api="fecha_emision"),
        attr(18, "cdr_estado", "VARCHAR", "50", json_api="cdr_estado"),
        attr(19, "pdf_url", "VARCHAR", "500", json_api="pdf_url"),
        attr(20, "cdr_zip_url", "VARCHAR", "500", json_api="cdr_zip_url"),
        attr(21, "doc_afectado_tipo", "VARCHAR", "20", json_api="documento_afectado.tipo"),
        attr(22, "doc_afectado_serie", "VARCHAR", "10", json_api="documento_afectado.serie"),
        attr(23, "doc_afectado_numero", "VARCHAR", "20", json_api="documento_afectado.numero"),
        attr(24, "doc_afectado_fecha", "TIMESTAMPTZ", json_api="documento_afectado.fecha_emision"),
        attr(25, "created_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(26, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
        attr(27, "—", "—", unico="Sí", notas="UNIQUE (company_ruc, direccion, tipo, serie, numero)"),
    ],
    "comprobante_linea": [
        attr(1, "id", "UUID", pk="PK", nullable="No", json_api="id"),
        attr(2, "comprobante_id", "UUID", fk="comprobante.id", nullable="No", indice="idx_comp_linea"),
        attr(3, "orden", "SMALLINT", nullable="No", default="1"),
        attr(4, "catalog_item_id", "UUID", fk="catalog_item.id", json_api="catalog_item_id"),
        attr(5, "descripcion", "VARCHAR", "500", nullable="No", json_api="descripcion", desc="Snapshot texto"),
        attr(6, "cantidad", "DECIMAL", "18,4", nullable="No", json_api="cantidad"),
        attr(7, "unidad", "VARCHAR", "3", nullable="No", json_api="unidad"),
        attr(8, "precio_unitario", "DECIMAL", "18,4", nullable="No", json_api="precio_unitario"),
        attr(9, "afectacion_igv", "CHAR", "2", nullable="No", default="10", json_api="afectacion_igv"),
        attr(10, "subtotal", "DECIMAL", "18,2", nullable="No", json_api="subtotal"),
        attr(11, "igv", "DECIMAL", "18,2", nullable="No", json_api="igv"),
        attr(12, "total", "DECIMAL", "18,2", nullable="No", json_api="total"),
    ],
    "comprobante_linea_serie": [
        attr(1, "comprobante_linea_id", "UUID", pk="PK", fk="comprobante_linea.id", nullable="No"),
        attr(2, "producto_serie_id", "UUID", pk="PK", fk="producto_serie.id", nullable="No"),
    ],
    "comprobante_factura_vinculada": [
        attr(1, "comprobante_guia_id", "UUID", pk="PK", fk="comprobante.id", nullable="No", desc="Guía de emisión"),
        attr(2, "factura_tipo", "VARCHAR", "20", pk="PK", nullable="No", json_api="facturas[].tipo"),
        attr(3, "factura_serie", "VARCHAR", "10", pk="PK", nullable="No", json_api="facturas[].serie"),
        attr(4, "factura_numero", "VARCHAR", "20", pk="PK", nullable="No", json_api="facturas[].numero"),
        attr(5, "factura_fecha_emision", "TIMESTAMPTZ", json_api="facturas[].fecha_emision"),
        attr(6, "factura_comprobante_id", "UUID", fk="comprobante.id", desc="FK opcional si existe en BD"),
    ],
    "secuencia_documento": [
        attr(1, "company_ruc", "VARCHAR", "11", pk="PK", fk="empresa.ruc", nullable="No"),
        attr(2, "tipo", "VARCHAR", "30", pk="PK", nullable="No", desc="FACTURA_F001, ENTREGA, …"),
        attr(3, "serie", "VARCHAR", "10", pk="PK", nullable="No"),
        attr(4, "ultimo_numero", "INTEGER", nullable="No", default="0"),
        attr(5, "updated_at", "TIMESTAMPTZ", nullable="No", default="now()"),
    ],
    "ENUMS": [
        attr(1, "ComprobanteEstado", "—", desc="BORRADOR, ENVIADO, ACEPTADO, RECHAZADO, ANULADO"),
        attr(2, "ComprobanteTipo", "—", desc="FACTURA, BOLETA, NOTA_CREDITO, NOTA_DEBITO, GUIA_EMISION"),
        attr(3, "MovimientoTipo", "—", desc="ENTRADA, SALIDA, AJUSTE"),
        attr(4, "MovimientoEstado", "—", desc="BORRADOR, DESPACHADA, ANULADA"),
        attr(5, "ProductoSerieEstado", "—", desc="DISPONIBLE, RESERVADO, VENDIDO, ENTREGADO, BAJA"),
        attr(6, "CatalogItemKind", "—", desc="PRODUCT, SERVICE"),
        attr(7, "EstadoUsuario", "—", desc="ACTIVO, DISABLED, DELETED"),
        attr(8, "BusinessTemplate", "—", desc="GENERAL, RETAIL, SERVICIOS"),
        attr(9, "TipoDocSUNAT", "—", desc="1=DNI, 6=RUC"),
        attr(10, "UnidadSUNAT", "—", desc="NIU, MTR, KGM, LTR, ZZ (servicios)"),
    ],
    "API_RUTAS": [
        attr(1, "POST", "/auth/login", desc="LoginRequest: email, pin"),
        attr(2, "GET", "/empresas/{ruc}/clientes", desc="List<Cliente>"),
        attr(3, "POST", "/empresas/{ruc}/clientes", desc="CrearClienteRequest"),
        attr(4, "GET", "/empresas/{ruc}/catalogo?almacen_id=", desc="List<CatalogItem>"),
        attr(5, "POST", "/empresas/{ruc}/comprobantes", desc="EmitirComprobanteRequest"),
        attr(6, "GET", "/empresas/{ruc}/comprobantes", desc="Ventas emitidas"),
        attr(7, "GET", "/empresas/{ruc}/compras", desc="Comprobantes RECIBIDO"),
        attr(8, "GET/POST", "/empresas/{ruc}/almacenes", desc="Almacen"),
        attr(9, "GET", "/empresas/{ruc}/catalogo/{id}/series-disponibles", desc="ProductoSerie"),
        attr(10, "POST", "/empresas/{ruc}/inventario/entradas", desc="RegistrarEntradaRequest"),
        attr(11, "GET", "/empresas/{ruc}/inventario/movimientos?tipo=", desc="Kardex"),
        attr(12, "GET/POST", "/empresas/{ruc}/entregas", desc="Salidas (Movimiento SALIDA)"),
    ],
    "VISTAS_API": [
        attr(1, "EventoVidaProducto", "—", desc="Vista sobre movimiento+lineas; no tabla", notas="historial / GET futuro"),
        attr(2, "stock_actual en CatalogItem", "—", desc="JOIN stock_almacen o COUNT series", notas="query ?almacen_id="),
    ],
}


def style_header(ws, row=1):
    for col in range(1, len(COLS) + 1):
        cell = ws.cell(row=row, column=col, value=COLS[col - 1])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.row_dimensions[row].height = 28


def write_table_sheet(wb, name, rows):
    title = name[:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=title)
    ws["A1"] = f"Tabla: {name}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws["A2"] = "Alineado con DTOs FactApp (app mobile)"
    ws["A2"].font = SUB_FONT
    style_header(ws, row=4)
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    widths = [6, 28, 14, 10, 6, 22, 10, 14, 8, 18, 22, 40, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_index_sheet(wb):
    ws = wb.create_sheet(title="00_INDICE", index=0)
    ws["A1"] = "Modelo relacional — Backend FactApp"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Hoja"
    ws["B3"] = "Descripción"
    ws["A3"].font = ws["B3"].font = Font(bold=True)
    descriptions = {
        "empresa": "Tenant / emisor (RUC)",
        "usuario": "Cuentas y autenticación",
        "usuario_empresa": "Multi-empresa por usuario",
        "sesion_refresh": "Refresh tokens",
        "cliente": "Clientes receptoras",
        "catalog_item": "Productos y servicios",
        "almacen": "Bodegas",
        "stock_almacen": "Stock por almacén (no serializado)",
        "producto_serie": "Unidades serializadas",
        "movimiento": "Kardex cabecera (entrada/salida/ajuste)",
        "movimiento_linea": "Renglones kardex",
        "movimiento_linea_serie": "Series por línea de movimiento",
        "comprobante": "Ventas y compras electrónicas",
        "comprobante_linea": "Detalle comprobante (snapshot)",
        "comprobante_linea_serie": "Series facturadas",
        "comprobante_factura_vinculada": "Guía → facturas",
        "secuencia_documento": "Correlativos serie/número",
        "ENUMS": "Valores enumerados",
        "API_RUTAS": "Endpoints Retrofit",
        "VISTAS_API": "DTOs calculados (no persistir)",
    }
    row = 4
    for sheet_name in wb.sheetnames:
        if sheet_name == "00_INDICE":
            continue
        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=descriptions.get(sheet_name, ""))
        row += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws["A" + str(row + 2)] = f"Generado: {Path(__file__).name}"
    ws["A" + str(row + 2)].font = SUB_FONT


def main():
    wb = Workbook()
    wb.remove(wb.active)
    order = [
        "empresa",
        "usuario",
        "usuario_empresa",
        "sesion_refresh",
        "cliente",
        "catalog_item",
        "almacen",
        "stock_almacen",
        "producto_serie",
        "movimiento",
        "movimiento_linea",
        "movimiento_linea_serie",
        "comprobante",
        "comprobante_linea",
        "comprobante_linea_serie",
        "comprobante_factura_vinculada",
        "secuencia_documento",
        "ENUMS",
        "API_RUTAS",
        "VISTAS_API",
    ]
    for name in order:
        write_table_sheet(wb, name, TABLES[name])
    write_index_sheet(wb)
    wb.save(OUTPUT)
    print(f"OK: {OUTPUT}")


if __name__ == "__main__":
    main()
